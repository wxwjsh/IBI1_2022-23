#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue May 16 17:54:44 2023

@author: jishuhan
"""

import xml.dom.minidom
from openpyxl import Workbook

# Find and load xml documents go_obo.xml'
xml_file_path =r'/Users/jishuhan/Desktop/go_obo.xml'
dom = xml.dom.minidom.parse(xml_file_path)

# find <term> elements
terms= dom.getElementsByTagName('term')

# Create an Excel that can be filled
#Learn from the Internet “how to make a excal"
workbook = Workbook()
sheet = workbook.active

# create a sheet with four column named 'GO ID', 'Term Name', 'Definition' and 'Number of Child Nodes'.
sheet['A1'] = 'GO ID'
sheet['B1'] = 'Term Name'
sheet['C1'] = 'Definition'
sheet['D1'] = 'Number of Child Nodes'

# Fill in the data from the second row
row_index = 2

#Inside the count_child_nodes() function, it iterates over all the <term> elements and checks if the <is_a> 
#element within the term has a value that matches the go_id of the input term. If there is a match, it increments 
#the count variable and recursively calls the count_child_nodes() function for the child term identified by the <id> element.
#This process continues until all child nodes have been counted.
#Learn from the Internet 
def count_child_nodes(term_id):
    count = 0
    for term_element in terms:
        is_a_elements = term_element.getElementsByTagName('is_a')
        for is_a_element in is_a_elements:
            if is_a_element.firstChild.data == term_id:
                count += 1
                count += count_child_nodes(term_element.getElementsByTagName('id')[0].firstChild.data)
    return count


# for all <term> elements, find <defstr> elements
for term_element in terms:
    defstr = term_element.getElementsByTagName('defstr')[0]
    #find autophagosome in <defstr>
    #firstChild means: gets the first child of the ID element, which may be a text node containing the term ID of the child node.
    #find the <id>, <name>, and <defstr> elements
    if 'autophagosome' in defstr.firstChild.data:
        id = term_element.getElementsByTagName('id')[0]
        name = term_element.getElementsByTagName('name')[0]
        go_id = id.firstChild.data
        term_name = name.firstChild.data
        defi = defstr.firstChild.data
        
        # Count the number of child nodes
        num = count_child_nodes(go_id)
               
        #Writing data to excal tables
        sheet.cell(row=row_index, column=1).value = go_id
        sheet.cell(row=row_index, column=2).value = term_name
        sheet.cell(row=row_index, column=3).value = defi
        sheet.cell(row=row_index, column=4).value = num
        row_index += 1

#save the sheet
excel_file_path = r'/Users/jishuhan/Desktop/autophagosome.xlsx'
workbook.save(excel_file_path)